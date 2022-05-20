import openpyxl
import os


# first need to change os directory to excel sheet path

#after that
workbook = openpyxl.load_workbook('Book1.xlsx')

print(type(workbook))
sheet = workbook.sheetnames
print(sheet)
sheet1 = workbook['Sheet1']
print(sheet1['A1'])
cell = sheet1._get_cell(1,2).value

print(str(cell))
print(sheet1.cell(1,1).value)

for i in range(1,8):
    print(i,' ' + sheet1.cell(row=i,column=2).value)


# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).
# openpyxl.load_workbook(filename) returns a Workbook object.
# get_sheet_names() and get_sheet_by_name() help get Worksheet objects.
# The square brackets in sheet[‘A1'] get Cell objects.
# Cell objects have a "value" member variable with the content of that cell.
# The cell() method also returns a Cell object from a sheet.