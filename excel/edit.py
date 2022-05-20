import openpyxl

#create workbook object
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb['Sheet']
print(sheet.cell(1,1).value)

sheet.cell(1,1, 100)
print(sheet.cell(1,1).value)

sheet.cell(1,2, "Hello")
print(sheet.cell(1,2).value)

# import os 
# print(os.getcwd())
# wb.save('example_edit.xlsx')

sheet2 = wb.create_sheet()
print(wb.sheetnames)
sheet2.title = 'Just new sheet name'
print(wb.sheetnames)

#can position sheet
sheet3 = wb.create_sheet(index=0, title='new sheet but at the first')
print(wb.sheetnames)
